#!/usr/bin/env python3
"""生成三库（量表库/归因库/工具库）业务填写模板。

模板是业务与引擎之间的契约：教研组按此填写，导入器按此解析。
改字段就是改契约，必须同步 scripts/import-business-data 与 shared/contracts.ts。

用法: python3 scripts/generate-library-templates.py
输出: business-libraries/templates/三库填写模板_v1.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "business-libraries" / "templates" / "三库填写模板_v1.xlsx"

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
REQ_FILL = PatternFill("solid", fgColor="FFF2CC")   # 必填列
EXAMPLE_FONT = Font(name=FONT, italic=True, color="808080", size=10)
BODY_FONT = Font(name=FONT, size=10)
NOTE_FONT = Font(name=FONT, size=10, color="C00000")
TITLE_FONT = Font(name=FONT, bold=True, size=13)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---- 枚举 ----
MODULES = [
    ("self_growth", "自我成长赋能"),
    ("class_system", "班级系统建设"),
    ("home_school", "家校沟通合作"),
    ("student_case", "学生个体问题"),
    ("learning_problem", "学生学习问题"),
]
DIVISIONS = [("all", "全学部"), ("primary", "小学部"), ("junior", "初中部"),
             ("senior", "高中部"), ("repeat", "复读部")]
RESPONDENTS = [("teacher", "班主任本人"), ("student", "学生"),
               ("guardian", "家长"), ("class", "班级整体")]
SEVERITIES = [("low", "轻度"), ("medium", "中度"), ("high", "重度"), ("crisis", "危机")]
SENSITIVITY = [("internal", "内部限阅"), ("sensitive", "敏感"), ("highly_sensitive", "高度敏感")]
SOURCE_KIND = [("proprietary", "六力自有"), ("external", "外部量表·需授权"), ("adapted", "改编自外部")]
RISK_LEVELS = [("red", "红线·立即熔断"), ("orange", "橙·4小时内响应"),
               ("yellow", "黄·关注"), ("none", "无风险标记")]
YES_NO = ["是", "否"]


def enum_codes(pairs):
    return [c for c, _ in pairs]


def style_header(ws, headers, required):
    for idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        width = 34 if name in required and "步骤" in name else max(12, min(30, len(name) * 2.6))
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"


def write_example(ws, row_idx, values):
    for idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=idx, value=value)
        cell.font = EXAMPLE_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER


def add_list_validation(ws, col_letter, options, last_row=400):
    """短枚举用内联列表；避免跨表引用在部分 Excel 版本下失效。"""
    joined = ",".join(options)
    if len(joined) > 240:
        return
    dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True, showDropDown=False)
    dv.error = "请从下拉列表中选择"
    dv.errorTitle = "取值不在枚举内"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def sheet_readme(wb):
    ws = wb.create_sheet("① 填写说明")
    ws.column_dimensions["A"].width = 100
    lines = [
        ("三库填写模板 v1", TITLE_FONT),
        ("", None),
        ("这三个库是整个系统的主线：量表库决定问什么，归因库决定怎么判，工具库决定给什么。", BODY_FONT),
        ("三者靠「编码」互相咬合，编码错了系统就串不起来，因此编码列请务必按规则填写。", BODY_FONT),
        ("", None),
        ("【颜色约定】", Font(name=FONT, bold=True, size=11)),
        ("· 深蓝表头 = 列名，请勿修改、勿调整列顺序、勿插入新列。", BODY_FONT),
        ("· 每张表第 2 行是灰色斜体示例行，填写前请整行删除。", BODY_FONT),
        ("· 列名带 * 号 = 必填，留空会导致导入失败。", BODY_FONT),
        ("", None),
        ("【三库如何咬合】", Font(name=FONT, bold=True, size=11)),
        ("1. 量表-题目.题号  →  被 归因-分级规则.触发条件 直接引用。改题号等于改规则。", BODY_FONT),
        ("2. 归因-分级规则.工具标签  →  匹配 工具-处方总表.工具标签。两边写法必须完全一致。", BODY_FONT),
        ("3. 关键词-路由.所属模块  →  决定 AI 入口把教师的自然语言分流到哪个模块。", BODY_FONT),
        ("", None),
        ("【三条硬规则，请转告填写人】", Font(name=FONT, bold=True, size=11)),
        ("① 题号只能用英文字母、数字、下划线（如 q1、EE_3）。用中文或空格会导致规则引擎无法引用。", NOTE_FONT),
        ("② 每个模块的「归因-红线熔断」至少要有一条。没有红线的模块，系统会拒绝发布——"
         "因为那等于关掉了这个模块的危机熔断。", NOTE_FONT),
        ("③ 外部量表（如 SDQ）必须在「来源属性」选 external 并写明授权情况，否则不予发布。", NOTE_FONT),
        ("", None),
        ("【填写顺序建议】", Font(name=FONT, bold=True, size=11)),
        ("枚举字典（先看）→ 量表-选项组 → 量表-清单 → 量表-题目 → 归因-分级规则 → "
         "归因-红线熔断 → 工具-处方总表 → 工具-禁忌规则 → 关键词-路由", BODY_FONT),
        ("", None),
        ("【交回后会发生什么】", Font(name=FONT, bold=True, size=11)),
        ("导入器会逐行校验并出一份核对清单：编码是否重复、题号是否被规则引用但不存在、"
         "工具标签是否有归因规则指向、红线是否齐全。有错会指到具体行号，不会静默通过。", BODY_FONT),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    return ws


def sheet_enums(wb):
    ws = wb.create_sheet("② 枚举字典")
    headers = ["枚举类别", "取值（填这个）", "中文含义", "用在哪些列"]
    style_header(ws, headers, set())
    rows = []
    for code, label in MODULES:
        rows.append(("所属模块", code, label, "量表-清单 / 归因-分级规则 / 工具-处方总表 / 关键词-路由"))
    for code, label in DIVISIONS:
        rows.append(("适用学部", code, label, "量表-清单.适用学部 / 工具-处方总表.适用学部"))
    for code, label in RESPONDENTS:
        rows.append(("施测/适用对象", code, label, "量表-清单.施测对象 / 工具-处方总表.适用对象"))
    for code, label in SEVERITIES:
        rows.append(("严重度", code, label, "工具-处方总表.严重度"))
    for code, label in SENSITIVITY:
        rows.append(("数据敏感级", code, label, "量表-清单.数据敏感级"))
    for code, label in SOURCE_KIND:
        rows.append(("来源属性", code, label, "量表-清单.来源属性"))
    for code, label in RISK_LEVELS:
        rows.append(("风险等级", code, label, "关键词-路由.风险等级"))
    for r, values in enumerate(rows, start=2):
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 62
    return ws


def sheet_condition_guide(wb):
    ws = wb.create_sheet("⑤a 归因-条件写法")
    headers = ["你要表达的意思", "就这样写", "示例", "说明"]
    style_header(ws, headers, set())
    rows = [
        ("所有题目的总分", "总分", "总分 >= 20", "反向计分题已自动折算后再求和"),
        ("得分最高的那一题", "最高项", "最高项 >= 4", "对应「任一单项达到 4 分」"),
        ("得分最低的那一题", "最低项", "最低项 <= 2", ""),
        ("所有题目的平均分", "平均分", "平均分 < 2.2", ""),
        ("某一题的折算得分", "题[题号]", "题[q1] >= 4", "反向题取折算后的分；题号来自「量表-题目」"),
        ("某一题的原始作答", "原始分[题号]", "原始分[q3] <= 2", "不做反向折算，按教师实际勾选的分值"),
        ("某个维度的均值", "维度[维度名]", "维度[情绪状态] >= 3.5", "维度名来自「量表-题目.维度」，必须完全一致"),
        ("得分最高的维度是哪个", "主导维度", "主导维度 = 情绪", "用于「五类十五型」定主类"),
        ("得分最低的维度是哪个", "短板维度", "短板维度 = 关系", "用于「五系统」定短板"),
        ("同时满足两个条件", "且", "题[q1] >= 4 且 题[q3] >= 4", "六色预警的红色条件就是这样"),
        ("满足任一条件", "或", "总分 >= 20 或 最高项 >= 4", ""),
        ("连续低意义感的次数", "连续低意义感次数", "连续低意义感次数 >= 3", "系统按历史评估自动带入，业务只需引用"),
        ("兜底（什么都不满足时）", "（留空）", "", "每个模块必须且只能有一行留空，优先级填最大"),
    ]
    for r, values in enumerate(rows, start=2):
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in zip("ABCD", (26, 22, 34, 46)):
        ws.column_dimensions[col].width = width
    note = ws.cell(row=len(rows) + 3, column=1,
                   value="比较符号只支持： >=   >   <=   <   =   ≠ 。不支持加减乘除，"
                         "需要算式请改写成多行规则，或联系技术方增加计算变量。")
    note.font = NOTE_FONT
    note.alignment = Alignment(vertical="top", wrap_text=True)
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_enums(wb)

    # ---- ③ 量表-清单 ----
    ws = wb.create_sheet("③ 量表-清单")
    headers = ["量表编码*", "量表名称*", "所属模块*", "适用学部*", "施测对象*", "施测形式",
               "预计用时分钟*", "使用时机", "重评间隔天数", "责任角色", "数据敏感级*",
               "来源属性*", "外部授权说明", "手册出处*", "版本*", "量表说明"]
    style_header(ws, headers, set(headers))
    write_example(ws, 2, [
        "SG_FIVE_Q", "班主任状态五问", "self_growth", "all", "teacher", "自评问卷", 3,
        "每月一次", 30, "班主任本人", "highly_sensitive", "proprietary", "",
        "技术理论手册V5 第3章", "1.0.0", "回顾最近一周状态，产出六色提示"])
    add_list_validation(ws, "C", enum_codes(MODULES))
    add_list_validation(ws, "D", enum_codes(DIVISIONS))
    add_list_validation(ws, "E", enum_codes(RESPONDENTS))
    add_list_validation(ws, "K", enum_codes(SENSITIVITY))
    add_list_validation(ws, "L", enum_codes(SOURCE_KIND))

    # ---- ④ 量表-题目 ----
    ws = wb.create_sheet("④ 量表-题目")
    headers = ["量表编码*", "题号*", "维度*", "子维度", "题干*", "选项组编码*",
               "反向计分*", "权重", "答题提示"]
    style_header(ws, headers, set(headers))
    write_example(ws, 2, [
        "SG_FIVE_Q", "q1", "情绪状态", "", "这一周，我有多少时间感到身心疲惫、难以恢复？",
        "FREQ_5", "否", "", "按最近 7 天的实际感受作答"])
    write_example(ws, 3, [
        "SG_FIVE_Q", "q3", "意义感知", "", "这一周，有多少次我觉得「当班主任是值得的」？",
        "FREQ_5", "是", "", "这是反向题：选得越高代表状态越好"])
    add_list_validation(ws, "G", YES_NO)
    note = ws.cell(row=5, column=1, value="题号只能用英文字母、数字、下划线。归因规则会用「题[q1]」的形式直接引用它。")
    note.font = NOTE_FONT

    # ---- 量表-选项组 ----
    ws = wb.create_sheet("④b 量表-选项组")
    headers = ["选项组编码*", "选项顺序*", "选项文本*", "分值*"]
    style_header(ws, headers, set(headers))
    for i, (text, value) in enumerate(
            [("几乎没有", 1), ("很少", 2), ("有时", 3), ("经常", 4), ("几乎每天", 5)], start=1):
        write_example(ws, i + 1, ["FREQ_5", i, text, value])
    note = ws.cell(row=8, column=1, value="一组选项只需定义一次，多张量表可共用同一个选项组编码。")
    note.font = NOTE_FONT

    # ---- ⑤a 条件写法（先给业务看懂怎么写，再填规则表） ----
    sheet_condition_guide(wb)

    # ---- ⑤ 归因-分级规则 ----
    ws = wb.create_sheet("⑤b 归因-分级规则")
    headers = ["规则编码*", "所属模块*", "依据量表编码*", "优先级*", "触发条件",
               "命中等级*", "等级中文名*", "是否红线熔断*", "主归因*", "次归因",
               "工具标签*", "结果说明*", "手册出处"]
    style_header(ws, headers, set(headers))
    write_example(ws, 2, [
        "SG-RED-Q1-Q3", "self_growth", "SG_FIVE_Q", 10, "题[q1] >= 4 且 题[q3] >= 4",
        "red", "需立即关注", "是", "疲惫与意义感同时告急", "情绪耗竭",
        "self_growth,red", "疲惫与意义感风险同时处于高位，已暂停常规建议并转介。", "PRD 8.2.1"])
    write_example(ws, 3, [
        "SG-GREEN", "self_growth", "SG_FIVE_Q", 100, "",
        "green", "状态良好", "否", "状态稳定", "",
        "self_growth,green", "当前状态整体稳定，保持现有节奏即可。", "PRD 8.2.1"])
    add_list_validation(ws, "B", enum_codes(MODULES))
    add_list_validation(ws, "H", YES_NO)
    note = ws.cell(row=5, column=1,
                   value="优先级数字小的先判定，第一条命中即停止。每个模块必须且只能有一行「触发条件」留空作为兜底，"
                         "其优先级填最大值。「工具标签」是与工具库对接的唯一钥匙，多个标签用英文逗号分隔。")
    note.font = NOTE_FONT

    # ---- ⑥ 归因-红线熔断 ----
    ws = wb.create_sheet("⑥ 归因-红线熔断")
    headers = ["所属模块*", "红线条件*", "红线说明*", "处置要求*", "手册出处"]
    style_header(ws, headers, set(headers))
    write_example(ws, 2, [
        "self_growth", "题[q1] >= 4 且 题[q3] >= 4",
        "疲惫与意义感同时处于高位，属自我成长模块红线",
        "立即阻断常规建议输出，展示求助指引，生成转介工单通知心理专员", "PRD 8.3"])
    add_list_validation(ws, "A", enum_codes(MODULES))
    note = ws.cell(row=4, column=1,
                   value="五个模块每个至少一条。红线一旦命中，无论分级规则判成什么等级，系统都会强制熔断。"
                         "缺少红线的模块将被系统拒绝发布。")
    note.font = NOTE_FONT

    # ---- ⑦ 工具-处方总表 ----
    ws = wb.create_sheet("⑦ 工具-处方总表")
    headers = ["工具编码*", "工具名称*", "所属模块*", "工具形式*", "适用学部*", "适用对象*",
               "适用症状场景*", "严重度*", "对应归因*", "工具标签*", "作用维度",
               "操作步骤*", "关键话术", "预期效果*", "单次耗时", "疗程与频次",
               "重评间隔天数", "禁止事项*", "输出物", "协同工具编码", "手册出处*", "版本*"]
    style_header(ws, headers, set(headers))
    write_example(ws, 2, [
        "SG_RX_001", "三分钟补能法", "self_growth", "练习", "all", "teacher",
        "感到情绪即将失控、疲惫难以恢复时", "medium", "教师压力与资源失衡",
        "self_growth,orange,pressure", "情绪恢复",
        "1) 离开当前工作情境；2) 完成三轮缓慢呼吸；3) 命名此刻的情绪；4) 选一个最小行动",
        "现在先停三分钟，这三分钟只属于你自己。",
        "单次可下降 1-2 分主观压力值", "3 分钟", "每日 1-2 次，连续 7 天", 7,
        "不用于替代危机处置；出现自伤念头或持续失眠时须转介心理专员",
        "一周能量记录卡", "SG_RX_007", "个人成长处方库 P12", "1.0.0"])
    add_list_validation(ws, "C", enum_codes(MODULES))
    add_list_validation(ws, "E", enum_codes(DIVISIONS))
    add_list_validation(ws, "F", enum_codes(RESPONDENTS))
    add_list_validation(ws, "H", enum_codes(SEVERITIES))
    note = ws.cell(row=4, column=1,
                   value="「工具标签」必须能和「归因-分级规则.工具标签」对上，否则这条工具永远不会被推送。"
                         "多步骤用 1) 2) 3) 或换行分隔；多个标签用英文逗号分隔。")
    note.font = NOTE_FONT

    # ---- ⑧ 工具-禁忌规则 ----
    ws = wb.create_sheet("⑧ 工具-禁忌规则")
    headers = ["工具编码*", "禁忌条件*", "禁忌说明*", "依据"]
    style_header(ws, headers, set(headers))
    write_example(ws, 2, [
        "SG_RX_012", "互动模式 = A型",
        "A型（过度负责）教师禁止推送含「少付出／别那么负责」语义的内容", "PRD 8.2.1 依恋知情禁忌"])
    write_example(ws, 3, [
        "SG_RX_012", "依恋安全感 <= 5",
        "低依恋安全感时禁止任何激活恐惧的表述，强制先回应情感再给建议", "PRD 8.2.1"])
    note = ws.cell(row=5, column=1,
                   value="这张表和「处方总表.禁止事项」不同：那一列是给教师看的文字提醒，"
                         "这张表是给引擎执行的硬过滤——命中条件的工具会被直接从候选中剔除，教师根本看不到。"
                         "PRD 要求依恋知情禁忌必须在代码层拦截，靠的就是这张表。")
    note.font = NOTE_FONT

    # ---- ⑨ 关键词-路由 ----
    ws = wb.create_sheet("⑨ 关键词-路由")
    headers = ["关键词编码*", "核心触发词*", "扩展词与近义表达", "所属模块*", "风险等级*",
               "关联量表编码", "关联工具编码", "场景描述"]
    style_header(ws, headers, set(headers))
    write_example(ws, 2, [
        "KW_R01", "不想活", "想死；活着没意思；结束生命", "student_case", "red",
        "", "", "学生或教师在任意对话中表达自杀意念，任一命中即走危机流程"])
    write_example(ws, 3, [
        "KW_Y07", "家长投诉", "被家长投诉；家长找到校长；家长发朋友圈", "home_school", "yellow",
        "HS_QUICK", "HS_RX_003", "家长表达不满但未出现威胁行为"])
    add_list_validation(ws, "D", enum_codes(MODULES))
    add_list_validation(ws, "E", enum_codes(RISK_LEVELS))
    note = ws.cell(row=5, column=1,
                   value="这张表有两个用途：① 风险等级=red 的词直接接入危机熔断，替换目前硬编码在代码里的 5 条正则；"
                         "② 其余词用于 AI 入口的模块分流接地，让分流不再只靠大模型自由发挥。")
    note.font = NOTE_FONT

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"已生成: {OUT.relative_to(ROOT)}")
    print(f"工作表: {len(wb.sheetnames)} 张 -> {' / '.join(wb.sheetnames)}")


if __name__ == "__main__":
    build()
